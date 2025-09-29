"""Utility script to append test draft picks to the Excel canonical file.

Purpose:
  - Quickly verify live Numbers sync (ENABLE_LIVE_NUMBERS_SYNC) works while the
    corresponding .numbers document is open.
  - Safely append clearly marked test rows that you can later delete manually.

Usage (fish shell examples):
  set -x ENABLE_LIVE_NUMBERS_SYNC true
  python3 append_test_picks.py --players 3

  # Specify a custom Excel filename (must already exist from setup)
  python3 append_test_picks.py --file fantasy_draft_data.xlsx --players 5

  # Provide explicit player names (round/pick auto-assigned after existing rows)
  python3 append_test_picks.py --names "Connor McDavid" "Nathan MacKinnon" "Cale Makar"

Notes:
  - Script respects ENABLE_LIVE_NUMBERS_SYNC and NUMBERS_DOC_HINT env vars.
  - Only appends; no deletion/update.
  - Draft Results schema: round, pick, playerName, teamId, manager
"""

from __future__ import annotations

import argparse
import os
import sys
import logging
from datetime import datetime
from typing import List

from dotenv import load_dotenv  # type: ignore

# Load .env first so any ENABLE_LIVE_NUMBERS_SYNC there is respected
load_dotenv()

# Live sync always attempted on macOS exporter (flag no longer required).

import platform
if platform.system() == 'Darwin':
    from macos.exporter_macos import MacOSDraftExporter as DraftExporter  # type: ignore
else:
    from windows.exporter_xlsx import XlsxDraftExporter as DraftExporter  # type: ignore


LOGGER = logging.getLogger("append_test_picks")
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s - %(levelname)s - %(message)s",
)

DEFAULT_PLAYER_POOL = [
    "Connor McDavid",
    "Nathan MacKinnon",
    "Cale Makar",
    "Leon Draisaitl",
    "David Pastrnak",
    "Nikita Kucherov",
    "Mikko Rantanen",
    "Auston Matthews",
    "Jack Hughes",
    "Matthew Tkachuk",
]


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Append test draft picks to Draft Results sheet")
    parser.add_argument(
        "--file",
        default=os.getenv("EXCEL_FILENAME", "fantasy_draft_data.xlsx"),
        help="Excel canonical filename (must already exist)",
    )
    parser.add_argument(
        "--players",
        type=int,
        default=2,
        help="Number of sample players to append (ignored if --names provided)",
    )
    parser.add_argument(
        "--names",
        nargs="*",
        help="Explicit player names (override --players count)",
    )
    parser.add_argument(
        "--manager",
        default="Test Manager",
        help="Manager value to write in manager column",
    )
    parser.add_argument(
        "--team-id",
        default="teamTEST",
        help="Team ID value to write in teamId column",
    )
    parser.add_argument(
        "--round",
        type=int,
        default=None,
        help="Force round number (otherwise auto-detect / increment)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print planned rows but do not write anything",
    )
    return parser.parse_args(argv)


def determine_next_round_and_pick(xlsx_filename: str) -> tuple[int, int]:
    """Inspect existing Draft Results sheet to find next pick position.

    Returns (round, next_pick_number). If no data -> (1,1).
    """
    from openpyxl import load_workbook  # local import to avoid global dependency at import time
    try:
        wb = load_workbook(xlsx_filename)
        if "Draft Results" not in wb.sheetnames:
            return 1, 1
        ws = wb["Draft Results"]
        last_round = 1
        last_pick = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or (row[0] is None and row[1] is None):
                continue
            try:
                rnd = int(row[0]) if row[0] is not None else None
                pck = int(row[1]) if row[1] is not None else None
            except (TypeError, ValueError):
                continue
            if rnd is not None and pck is not None:
                if rnd > last_round or (rnd == last_round and pck > last_pick):
                    last_round, last_pick = rnd, pck
        return last_round, last_pick + 1
    except Exception:
        return 1, 1


def build_test_rows(
    exporter: DraftExporter,
    names: List[str],
    team_id: str,
    manager: str,
    forced_round: int | None,
) -> List[List[str]]:
    rows: List[List[str]] = []
    current_round, next_pick = determine_next_round_and_pick(exporter.filename)
    if forced_round is not None:
        current_round = forced_round
    for name in names:
        rows.append([
            current_round,
            next_pick,
            name,
            team_id,
            manager,
        ])
        next_pick += 1
    return rows


def main(argv: List[str]) -> int:
    args = parse_args(argv)

    # If user passed a .numbers file, map to the canonical .xlsx (NumbersExporter requires .xlsx)
    if args.file.lower().endswith('.numbers'):
        base = args.file[:-8] or 'fantasy_draft_data'
        mapped = base + '.xlsx'
        LOGGER.warning("Provided file '%s' is a .numbers package; using canonical Excel '%s' for writes.", args.file, mapped)
        args.file = mapped

    if not args.file.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        # Force .xlsx extension if unspecified/unsupported
        if '.' not in os.path.basename(args.file):
            args.file = args.file + '.xlsx'
            LOGGER.info("Normalized filename to '%s'", args.file)
        else:
            LOGGER.error("Unsupported extension for '%s'. Please use .xlsx", args.file)
            return 1

    if not os.path.exists(args.file):
        LOGGER.error("Excel file '%s' not found. Run setup first.", args.file)
        return 1

    exporter = DraftExporter(args.file)

    if args.names:
        chosen = args.names
    else:
        if args.players <= 0:
            LOGGER.error("--players must be > 0 (or provide --names)")
            return 1
        chosen = DEFAULT_PLAYER_POOL[: args.players]
        if len(chosen) < args.players:
            LOGGER.warning("Requested %d players but only %d available in pool", args.players, len(chosen))

    rows = build_test_rows(
        exporter=exporter,
        names=chosen,
        team_id=args.team_id,
        manager=args.manager,
        forced_round=args.round,
    )

    LOGGER.info("Prepared %d test draft picks:", len(rows))
    for r in rows:
        LOGGER.info("  round=%s pick=%s player=%s", r[0], r[1], r[2])

    if args.dry_run:
        LOGGER.info("Dry run enabled; no rows appended.")
        return 0

    exporter.append_draft_results(rows)  # compatibility alias in both exporters
    exporter.add_timestamp()
    LOGGER.info("Done. If live sync enabled, check the open Numbers document.")
    LOGGER.info("(If Numbers document is open on macOS it should have updated)")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main(sys.argv[1:]))
