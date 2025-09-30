import pytest
from openpyxl import load_workbook


@pytest.fixture()
def exporter(tmp_path):
    from windows.xlsx_export import XlsxDraftExporter
    return XlsxDraftExporter(str(tmp_path / 'extended.xlsx'))


def test_batch_append_picks(exporter):
    # Append in 3 batches and verify continuity
    exporter.append_picks([["1", 1, "p1", "t", "m"], ["1", 2, "p2", "t", "m"]])
    exporter.append_picks([["2", 3, "p3", "t", "m"]])
    exporter.append_picks([["2", 4, "p4", "t", "m"]])
    wb = load_workbook(exporter.filename, data_only=True)
    ws = wb['Draft Results']
    picks = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert picks == [1, 2, 3, 4]


def test_draft_board_formula_persistence(exporter):
    # Minimal pre-draft rows
    pda_rows = [
        ["k1", "Skater One", "NYR", "C", "10", "1", "100", "20", "18", "5", "2", "12", "90"],
        ["k2", "Goalie Two", "BOS", "G", "20", "2", "90", "15", "12", "15", "1", "18", "80"],
    ]
    exporter.update_draft_analysis_data(pda_rows)
    league_settings = {
        'stat_categories': [
            {'position_type': 'P', 'display_name': 'Goals', 'value': '4'},
            {'position_type': 'G', 'display_name': 'Wins', 'value': '5'},
        ]
    }
    exporter.setup_projection_sheets(league_settings)
    exporter.create_draft_board()

    # Re-open and verify formulas still present
    wb = load_workbook(exporter.filename, data_only=False)
    db = wb['Draft Board']
    assert db['A2'].value.startswith('=IFERROR(VLOOKUP(')
    assert db['F2'].value.startswith('=IF(')


@pytest.mark.slow
def test_projection_sheet_stress_formulas(exporter):
    # Build league settings with a couple of stats
    league_settings = {
        'stat_categories': [
            {'position_type': 'P', 'display_name': 'Goals', 'value': '4'},
            {'position_type': 'P', 'display_name': 'Assists', 'value': '3'},
        ]
    }
    exporter.setup_projection_sheets(league_settings)

    # Add many rows to skater projections to simulate user input, ensure TOTAL formula template exists deep.
    wb = load_workbook(exporter.filename)
    sk = wb['Skater Projections']
    # Row 2..1501 already prepped for formulas by exporter; verify formula cell at row ~1500
    formula = sk.cell(row=1500, column=4).value  # TOTAL column
    assert formula is not None and formula.startswith('=')
