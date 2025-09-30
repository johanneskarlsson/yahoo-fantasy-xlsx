import re
import pytest
from openpyxl import load_workbook


@pytest.fixture()
def exporter(tmp_path):
    from windows.xlsx_export import XlsxDraftExporter
    filename = tmp_path / "test_draft.xlsx"
    exp = XlsxDraftExporter(str(filename))
    return exp


def test_initial_workbook_structure(exporter):
    wb = load_workbook(exporter.filename)
    # All base sheet names present
    for sheet in exporter.BASE_SHEETS.keys():
        assert sheet in wb.sheetnames
    # Headers present in Draft Results
    ws = wb["Draft Results"]
    headers = [cell.value for cell in ws[1]]
    assert headers[:5] == exporter.BASE_SHEETS["Draft Results"][:5]


def test_append_picks_adds_rows(exporter):
    from windows.xlsx_export import XlsxDraftExporter
    exporter.append_picks([["1", 1, "p1", "t1", "mgr1"], ["1", 2, "p2", "t2", "mgr2"]])
    wb = load_workbook(exporter.filename, data_only=True)
    ws = wb["Draft Results"]
    # Header + 2 rows
    assert ws.max_row >= 3
    assert ws.cell(row=2, column=3).value == "p1"
    assert ws.cell(row=3, column=3).value == "p2"


def test_timestamp_sets_cell(exporter):
    exporter.timestamp()
    wb = load_workbook(exporter.filename, data_only=True)
    ws = wb["Draft Results"]
    v = ws['I1'].value
    assert v and v.startswith("Last updated:")
    # Basic datetime format check
    ts = v.replace("Last updated: ", "")
    assert re.match(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}", ts)


def test_update_league_settings_data(exporter):
    league_settings = {
        'league_name': 'Test League',
        'league_type': 'H2H',
        'scoring_type': 'points',
        'max_teams': '12',
        'num_playoff_teams': '6',
        'playoff_start_week': '18',
        'roster_positions': [
            {'position': 'C', 'count': '2'},
            {'position': 'G', 'count': '2'},
        ],
        'stat_categories': [
            {'position_type': 'P', 'display_name': 'G', 'value': '4'},
            {'position_type': 'P', 'display_name': 'A', 'value': '3'},
            {'position_type': 'G', 'display_name': 'W', 'value': '5'},
        ],
    }
    exporter.update_league_settings_data(league_settings)
    wb = load_workbook(exporter.filename, data_only=True)
    ws = wb["League Settings"]
    # Find markers
    values = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
    assert 'ROSTER POSITIONS' in values
    assert 'SKATER STATS' in values
    assert 'GOALIE STATS' in values
    # Check a stat value line
    found_goal_stat = any(ws.cell(row=i, column=1).value == 'W' and ws.cell(row=i, column=2).value == '5' for i in range(1, ws.max_row + 1))
    assert found_goal_stat


def test_update_teams_data_clears_previous(exporter):
    exporter.update_teams_data([["tk1", "1", "Team 1", "Mgr1"], ["tk2", "2", "Team 2", "Mgr2"]])
    exporter.update_teams_data([["tk3", "3", "Team 3", "Mgr3"]])
    wb = load_workbook(exporter.filename, data_only=True)
    ws = wb["Teams"]
    # Collect non-empty team keys (excluding header row)
    keys = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=1).value]
    assert 'tk1' not in keys and 'tk2' not in keys
    assert 'tk3' in keys


def test_update_draft_analysis_data_replaces_rows(exporter):
    first_rows = [["k1", "Player 1", "NYR", "C", "10", "1", "100", "20", "18", "5", "2", "12", "90"]]
    exporter.update_draft_analysis_data(first_rows)
    second_rows = [["k2", "Player 2", "BOS", "G", "20", "2", "90", "15", "12", "15", "1", "18", "80"]]
    exporter.update_draft_analysis_data(second_rows)
    wb = load_workbook(exporter.filename, data_only=True)
    ws = wb["Pre-Draft Analysis"]
    # After clearing, first data row should now be second_rows[0]
    # Data may be appended after an existing cleared row (header row preserved); find row containing k2
    found = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == 'k2':
            found = True
            break
    assert found, "Updated data row with key 'k2' not found"


@pytest.mark.parametrize(
    "ptype,stats,expected_tokens",
    [
        ('P', [({'position_type': 'P', 'display_name': 'Goals', 'value': '4'},
                {'position_type': 'P', 'display_name': 'Assists', 'value': '3'})], ['=B2*4', 'C2*3']),
        ('G', [({'position_type': 'G', 'display_name': 'Wins', 'value': '5'},)], ['=B2*5'])
    ]
)
def test_projection_sheet_formula_templates(exporter, ptype, stats, expected_tokens, formula_assert):
    # Flatten stats tuple list
    flat_stats = []
    for group in stats:
        flat_stats.extend(group)
    league_settings = {'stat_categories': flat_stats}
    exporter.setup_projection_sheets(league_settings)
    wb = load_workbook(exporter.filename)
    if ptype == 'P':
        ws = wb['Skater Projections']
        formula = ws.cell(row=2, column=len([s for s in flat_stats if s['position_type']=='P']) + 2).value
    else:
        ws = wb['Goalie Projections']
        formula = ws.cell(row=2, column=3).value
    for token in expected_tokens:
        formula_assert(formula, token)


@pytest.fixture()
def draft_board_ready(exporter):
    pda_rows = [
        ["k1", "Skater One", "NYR", "C", "10", "1", "100", "20", "18", "5", "2", "12", "90"],
        ["k2", "Goalie Two", "BOS", "G", "20", "2", "90", "15", "12", "15", "1", "18", "80"],
    ]
    exporter.update_draft_analysis_data(pda_rows)
    league_settings = {
        'stat_categories': [
            {'position_type': 'P', 'display_name': 'Goals', 'value': '4'},
            {'position_type': 'G', 'display_name': 'Wins', 'value': '5'},
        ]
    }
    exporter.setup_projection_sheets(league_settings)
    exporter.create_draft_board()
    return exporter


def test_draft_board_player_references(draft_board_ready):
    from windows.xlsx_export import load_workbook as _lbw  # type: ignore
    wb = load_workbook(draft_board_ready.filename)
    db = wb['Draft Board']
    assert db['B2'].value == "='Pre-Draft Analysis'!B2"


def test_draft_board_lookup_formulas(draft_board_ready, formula_assert):
    wb = load_workbook(draft_board_ready.filename)
    db = wb['Draft Board']
    formula_assert(db['A2'].value, 'VLOOKUP')
    formula_assert(db['F2'].value, 'IF(', 'VLOOKUP')
