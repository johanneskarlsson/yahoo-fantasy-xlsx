import pytest
from openpyxl import load_workbook


def test_windows_exporter_has_adapters(tmp_path):
    from windows.xlsx_export import XlsxDraftExporter
    exp = XlsxDraftExporter(str(tmp_path / 'file.xlsx'))
    assert hasattr(exp, 'append_draft_results')
    assert hasattr(exp, 'add_timestamp')


def test_draft_monitor_main_runs_one_cycle_without_monkeypatch(tmp_path, monkeypatch, capsys):
    """Integration-lite: Run windows draft_monitor.main for a single cycle using adapter methods."""
    # Force filename to temporary path
    monkeypatch.setenv('FILENAME', str(tmp_path / 'mon.xlsx'))
    # Pretend we're on Windows so draft_monitor picks XlsxDraftExporter
    monkeypatch.setattr('platform.system', lambda: 'Windows')
    from importlib import reload
    from windows import draft_monitor as dm
    reload(dm)  # re-evaluate with new FILENAME

    class FakeAPI:
        def __init__(self):
            self.calls = 0
        def ensure_authenticated(self):
            return True
        def get_draft_results(self):
            if self.calls == 0:
                self.calls += 1
                return [
                    {"pick": {"#text": "1"}, "round": {"#text": "1"}, "team_key": {"#text": "T"}, "player_key": "p1"},
                ]
            raise KeyboardInterrupt

    monkeypatch.setattr(dm, 'api', FakeAPI())
    monkeypatch.setattr(dm, 'INTERVAL', 0)
    # speed time
    monkeypatch.setattr(dm, 'time', type('T', (), {'time': staticmethod(lambda: 0), 'sleep': staticmethod(lambda x: None)}))

    dm.main()
    out = capsys.readouterr().out
    assert 'Pick 1: p1' in out

    # Verify workbook updated
    wb = load_workbook(dm.exporter.filename)
    ws = wb['Draft Results']
    assert ws.max_row >= 2
    assert ws.cell(row=2, column=3).value == 'p1'