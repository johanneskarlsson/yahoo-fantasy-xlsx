import os
import logging
from openpyxl import Workbook, load_workbook
from datetime import datetime
from typing import List, Dict, Any
from abc import ABC, abstractmethod


class BaseDraftExporter(ABC):
    """Shared base class for platform-specific draft exporters."""

    BASE_SHEETS = {
        "Draft Board": ["draftedBy", "playerName", "team", "position", "averagePick", "projectedPoints"],
        "League Settings": ["setting", "value"],
        "Teams": ["teamKey", "teamId", "teamName", "manager"],
        "Draft Results": ["round", "pick", "playerName", "teamId", "manager"],
        "Pre-Draft Analysis": [
            "playerKey", "playerName", "team", "position", "averagePick", "averageRound", "percentDrafted",
            "projectedAuctionValue", "averageAuctionCost", "seasonRank", "positionRank", "preseasonAveragePick",
            "preseasonPercentDrafted"
        ]
    }

    def __init__(self, filename: str = "fantasy_draft_data.xlsx"):
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'
        self.filename = filename
        self.logger = logging.getLogger(__name__)
        if not os.path.exists(self.filename):
            self._create_full_base()
        else:
            self._verify_sheets()

    def _create_full_base(self):
        """Create a new workbook with all base sheets."""
        wb = Workbook()
        if wb.active.title == 'Sheet':
            wb.remove(wb.active)
        for sheet, headers in self.BASE_SHEETS.items():
            ws = wb.create_sheet(sheet)
            self._style_headers(ws, headers)
        wb.save(self.filename)
        self.logger.debug(f"Created new workbook {self.filename} with base sheets")

    def _verify_sheets(self):
        """Ensure all required sheets exist in the workbook."""
        wb = load_workbook(self.filename)
        changed = False
        for sheet, headers in self.BASE_SHEETS.items():
            if sheet not in wb.sheetnames:
                ws = wb.create_sheet(sheet)
                self._style_headers(ws, headers)
                changed = True
        if changed:
            wb.save(self.filename)

    def _style_headers(self, ws, headers):
        """Apply styling to header row."""
        try:
            from openpyxl.styles import Font, PatternFill
        except Exception:
            Font = None
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            if 'Font' in globals() and Font:
                cell.font = Font(bold=True)
                try:
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                except Exception:
                    pass

    def append_picks(self, rows: List[List[str]]):
        """Append new draft picks to Draft Results sheet."""
        if not rows:
            return
        wb = load_workbook(self.filename)
        ws = wb["Draft Results"]
        start = ws.max_row + 1
        for r_index, row in enumerate(rows, start=start):
            for c_index, val in enumerate(row, start=1):
                ws.cell(row=r_index, column=c_index, value=val)
        wb.save(self.filename)
        self.logger.debug(f"Added {len(rows)} picks (rows {start}-{start+len(rows)-1})")
        # Call platform-specific post-append hook
        self._post_append_hook(rows, start)

    def timestamp(self):
        """Add timestamp to Draft Results sheet."""
        wb = load_workbook(self.filename)
        ws = wb["Draft Results"]
        ws['I1'] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        wb.save(self.filename)

    # Legacy method aliases for compatibility
    def append_draft_results(self, rows):
        self.append_picks(rows)

    def add_timestamp(self):
        self.timestamp()

    def update_league_settings_data(self, league_settings: Dict[str, Any]):
        """Populate League Settings sheet with grouped sections."""
        try:
            wb = load_workbook(self.filename)
            if "League Settings" not in wb.sheetnames:
                return
            ws = wb["League Settings"]
            # Clear existing (keep headers row 1)
            for row in list(ws.iter_rows(min_row=2)):
                for cell in row:
                    cell.value = None

            rows = []
            rows.append(["League Name", league_settings.get('league_name', '')])
            rows.append(["League Type", league_settings.get('league_type', '')])
            rows.append(["Scoring Type", league_settings.get('scoring_type', '')])
            rows.append(["Max Teams", league_settings.get('max_teams', '')])
            rows.append(["Playoff Teams", league_settings.get('num_playoff_teams', '')])
            rows.append(["Playoff Start Week", league_settings.get('playoff_start_week', '')])
            rows.append(["", ""])  # spacer
            rows.append(["ROSTER POSITIONS", "COUNT"])
            for pos in league_settings.get('roster_positions', []):
                rows.append([pos.get('position', ''), pos.get('count', '')])
            rows.append(["", ""])  # spacer
            rows.append(["SKATER STATS", "VALUE"])
            for stat in league_settings.get('stat_categories', []):
                if stat.get('position_type') == 'P':
                    name = stat.get('display_name') or stat.get('name') or ''
                    rows.append([name, stat.get('value', '')])
            rows.append(["", ""])  # spacer
            rows.append(["GOALIE STATS", "VALUE"])
            for stat in league_settings.get('stat_categories', []):
                if stat.get('position_type') == 'G':
                    name = stat.get('display_name') or stat.get('name') or ''
                    rows.append([name, stat.get('value', '')])

            # Write rows
            for i, row in enumerate(rows, start=2):
                for j, val in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=val)
            wb.save(self.filename)
        except Exception as e:
            self.logger.debug(f"Failed to write league settings: {e}")

    def update_teams_data(self, teams_rows):
        """Write Teams sheet data."""
        if not teams_rows:
            return
        try:
            wb = load_workbook(self.filename)
            sheet = "Teams"
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in list(ws.iter_rows(min_row=2)):
                    for cell in row:
                        cell.value = None
            else:
                ws = wb.create_sheet(title=sheet)
                ws.append(["team_key", "team_id", "team_name", "manager"])
            for r in teams_rows:
                ws.append(r)
            wb.save(self.filename)
        except Exception as e:
            self.logger.debug(f"Failed to write teams data: {e}")

    def update_draft_analysis_data(self, players_rows):
        """Write Pre-Draft Analysis sheet data."""
        if not players_rows:
            return
        try:
            wb = load_workbook(self.filename)
            name = "Pre-Draft Analysis"
            if name not in wb.sheetnames:
                ws = wb.create_sheet(title=name)
                self._style_headers(ws, self.BASE_SHEETS[name])
            else:
                ws = wb[name]
                # Clear rows except header
                for row in list(ws.iter_rows(min_row=2)):
                    for cell in row:
                        cell.value = None
            for r in players_rows:
                if r:
                    ws.append(r)
            wb.save(self.filename)
        except Exception as e:
            self.logger.debug(f"Failed to write pre-draft analysis: {e}")

    def setup_projection_sheets(self, league_settings):
        """Create Skater/Goalie Projections sheets with TOTAL formulas."""
        try:
            wb = load_workbook(self.filename)
            # Determine stat names by position_type
            skater_stats, goalie_stats = [], []
            for stat in league_settings.get('stat_categories', []):
                name = stat.get('display_name') or stat.get('name') or ''
                ptype = stat.get('position_type')
                if ptype == 'P':
                    skater_stats.append(name)
                elif ptype == 'G':
                    goalie_stats.append(name)

            # Build sheets
            if skater_stats:
                if "Skater Projections" not in wb.sheetnames:
                    ws = wb.create_sheet("Skater Projections")
                    self._style_headers(ws, ["playerName"] + skater_stats + ["TOTAL"])
                else:
                    ws = wb["Skater Projections"]
                    for row in list(ws.iter_rows(min_row=2)):
                        for cell in row:
                            cell.value = None

            if goalie_stats:
                if "Goalie Projections" not in wb.sheetnames:
                    wg = wb.create_sheet("Goalie Projections")
                    self._style_headers(wg, ["playerName"] + goalie_stats + ["TOTAL"])
                else:
                    wg = wb["Goalie Projections"]
                    for row in list(wg.iter_rows(min_row=2)):
                        for cell in row:
                            cell.value = None

            wb.save(self.filename)
            # Add formula templates
            if skater_stats:
                self._setup_total_formulas("Skater Projections", skater_stats, league_settings)
            if goalie_stats:
                self._setup_total_formulas("Goalie Projections", goalie_stats, league_settings)
        except Exception as e:
            self.logger.debug(f"Failed to setup projection sheets: {e}")

    def _setup_total_formulas(self, sheet_name: str, stat_names, league_settings):
        """Set up TOTAL column formulas for projection sheets."""
        try:
            wb = load_workbook(self.filename)
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]
            ptype = 'G' if 'Goalie' in sheet_name else 'P'
            values = {}
            for stat in league_settings.get('stat_categories', []):
                if stat.get('position_type') == ptype:
                    name = stat.get('display_name') or stat.get('name') or ''
                    try:
                        values[name] = float(stat.get('value') or 0)
                    except Exception:
                        values[name] = 0
            TOTAL_col = len(stat_names) + 2  # playerName + stats + TOTAL
            for row_idx in range(2, 1502):
                parts = []
                for i, stat_name in enumerate(stat_names):
                    val = values.get(stat_name, 0)
                    if val:
                        col_letter = chr(66 + i)  # B onward
                        parts.append(f"{col_letter}{row_idx}*{val}")
                if parts:
                    ws.cell(row=row_idx, column=TOTAL_col, value="=" + "+".join(parts))
            wb.save(self.filename)
        except Exception as e:
            self.logger.debug(f"Failed to set total formulas for {sheet_name}: {e}")

    def create_draft_board(self):
        """Create Draft Board with formulas referencing other sheets."""
        try:
            wb = load_workbook(self.filename)
            if "Pre-Draft Analysis" not in wb.sheetnames:
                return
            analysis = wb["Pre-Draft Analysis"]
            if "Draft Board" not in wb.sheetnames:
                db = wb.create_sheet("Draft Board")
                self._style_headers(db, self.BASE_SHEETS["Draft Board"])
            else:
                db = wb["Draft Board"]
                # Clear existing rows (keep header)
                for row in list(db.iter_rows(min_row=2)):
                    for cell in row:
                        cell.value = None

            max_row = analysis.max_row
            # Populate with direct cell references
            for r in range(2, max_row + 1):
                if not analysis[f"A{r}"].value:
                    continue
                # draftedBy via Draft Results lookup (using player_key from Pre-Draft Analysis column A)
                db[f"A{r}"] = f"=IFERROR(VLOOKUP('Pre-Draft Analysis'!A{r},'Draft Results'!C:D,2,FALSE),\"\")"
                # Direct references for playerName, team, position, averagePick
                db[f"B{r}"] = f"='Pre-Draft Analysis'!B{r}"
                db[f"C{r}"] = f"='Pre-Draft Analysis'!C{r}"
                db[f"D{r}"] = f"='Pre-Draft Analysis'!D{r}"
                db[f"E{r}"] = f"='Pre-Draft Analysis'!E{r}"
                # projectedPoints: choose goalie vs skater projection VLOOKUP (using playerName)
                db[f"F{r}"] = (
                    f"=IF('Pre-Draft Analysis'!D{r}=\"G\"," \
                    f"IFERROR(VLOOKUP('Pre-Draft Analysis'!B{r},'Goalie Projections'!A:F,6,FALSE),\"\")," \
                    f"IFERROR(VLOOKUP('Pre-Draft Analysis'!B{r},'Skater Projections'!A:I,9,FALSE),\"\"))"
                )
            wb.save(self.filename)
        except Exception as e:
            self.logger.debug(f"Failed to create draft board: {e}")

    @abstractmethod
    def _post_append_hook(self, rows: List[List[str]], start_row: int):
        """Platform-specific hook called after appending picks. Override in subclasses."""
        pass